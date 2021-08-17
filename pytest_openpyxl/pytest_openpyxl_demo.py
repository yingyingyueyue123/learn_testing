from openpyxl import load_workbook, Workbook

if __name__ == "__main__":
    # 创建一个workbook

    file_name = r'test.xlsx'

    wb = Workbook()

    # 创建一个sheet，名为iTesting，把它插入到最前的位置

    wb.create_sheet('iTesting', 0)

    # 创建一个sheet，名为VIPTEST，把它插入index为1的位置

    wb.create_sheet('VIPTEST', 1)

    # 保存表格

    wb.save(file_name)

    # 读和写

    # 初始化表格

    wb2 = load_workbook(file_name)

    # 读，获取所有的sheet名称

    print(wb2.sheetnames)

    # 获取sheet名为iTesting的表格

    s = wb2['iTesting']

    # 将A1行的值设置为iTesting

    s['A1'] = 'iTesting'

    # 将第2行，第一列的值设置为1

    s.cell(row=2, column=1).value = 1

    # 打印第2行第一列单元格的值 --方法1

    print(s.cell(row=2, column=1).value)

    # 打印第2行第一列单元格的值 --方法2

    print(s['A2'].value)

    # 保存表格

    wb.save(file_name)
