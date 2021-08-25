# -*- coding: utf-8 -*-

import codecs

import json

import os

import time

import pytest

import yaml

from openpyxl import load_workbook

import pandas as pd


# 读取Yaml文件和Json文件

def read_data_from_json_yaml(data_file):
    return_value = []

    data_file_path = os.path.abspath(data_file)

    print(data_file_path)

    _is_yaml_file = data_file_path.endswith((".yml", ".yaml"))

    with codecs.open(data_file_path, 'r', 'utf-8') as f:

        # Load the data from YAML or JSON

        if _is_yaml_file:

            data = yaml.safe_load(f)

        else:

            data = json.load(f)

    for i, elem in enumerate(data):

        if isinstance(data, dict):

            key, value = elem, data[elem]

            if isinstance(value, dict):

                case_data = []

                for v in value.values():
                    case_data.append(v)

                return_value.append(tuple(case_data))

            else:

                return_value.append((value,))

    return return_value


# 读取Excel 文件 -- openpyxl

def read_data_from_excel(excel_file, sheet_name):
    return_value = []

    if not os.path.exists(excel_file):
        raise ValueError("File not exists")

    wb = load_workbook(excel_file)

    for s in wb.sheetnames:

        if s == sheet_name:

            sheet = wb[sheet_name]

            for row in sheet.rows:
                return_value.append([col.value for col in row])

    print(return_value)

    return return_value[1:]


# 读取Excel文件 -- Pandas

def read_data_from_pandas(excel_file, sheet_name):
    if not os.path.exists(excel_file):
        raise ValueError("File not exists")

    s = pd.ExcelFile(excel_file)

    df = s.parse(sheet_name)

    return df.values.tolist()


@pytest.mark.baidu
class TestBaidu:

    @pytest.mark.parametrize('search_string, expect_string', read_data_from_pandas(r'test_baidu_ddt.xlsx', 'iTesting'))
    def test_baidu_search(self, login, search_string, expect_string):
        driver, s, base_url = login

        driver.get(base_url + "/")

        driver.find_element_by_id("kw").send_keys(search_string)

        driver.find_element_by_id("su").click()

        time.sleep(2)

        search_results = driver.find_element_by_xpath('//*[@id="1"]/div/h3/a').get_attribute('innerHTML')

        print(search_results)

        assert (expect_string in search_results) is True


if __name__ == "__main__":
    pytest.main(['-s', '-v', 'test_baidu_ddt_pandas'])
